# -*- coding: utf-8 -*-
"""
bond_coupon_alert.py — 海外債配息「最晚買入日」提醒
====================================================
用法（獨立測試）：
    python bond_coupon_alert.py 08-14-2026-Bond_Pricing_Update_excel.xlsx

在龍蝦Bot裡：
    from bond_coupon_alert import build_alert_message
    msg = build_alert_message("path/to/latest.xlsx", today=date.today())

邏輯（依 Albert 規則）：
  1. 配息日 = 從「到期日」+「配息頻率」倒推（同一天號，每半年/每季/每年/每月）
  2. 最晚交割日 = 配息日的前 1 個營業日
  3. ISIN 開頭 US / CA → T+1；其他 (XS/AU/NZ/GB/…) → T+2
     最晚下單日 = 最晚交割日 再往前 N 個營業日
  4. 只列出「配息日落在 今天 ~ 今天+14天」的債券
     ※ 營業日只排除週六日，未排除台美假日
"""
import re
import sys
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

FREQ_MONTHS = {"每月": 1, "每季": 3, "每半年": 6, "每年": 12}
LOOKAHEAD_DAYS = 14

DISCLAIMER = (
    "\n⚠️ 已知限制\n"
    "1. 配息日是從「到期日＋配息頻率」倒推的，少數債券付息日與到期日不同號，請以實際配息日為主\n"
    "2. 營業日只避開週六日，未避開台美假日，假日前後請人工再確認\n"
    "3. 🔒專投＝限專業投資人（依報價檔分頁或備註判斷）；💎高資產＝高資產客戶專屬\n"
    "4. 本表僅列示配息時點資訊，非投資建議，亦非鼓勵於特定時點進場"
)

def is_bond_pricing_file(path, filename=""):
    """判斷上傳的 Excel 是不是總行的海外債報價檔（跟 ELN 檔區分開）"""
    if "bond" in str(filename).lower() and "pric" in str(filename).lower():
        return True
    try:
        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return any(("海外債券資訊" in n) or ("加碼標的" in n) for n in names)
    except Exception:
        return False

# ---------- 小工具 ----------
def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def settle_lag(isin):
    """US / CA 開頭 T+1，其他 T+2"""
    return 1 if str(isin).upper().startswith(("US", "CA")) else 2

def biz_days_before(d, n):
    """從 d 往前推 n 個營業日（只跳過週六日）"""
    cur = d
    while n > 0:
        cur -= timedelta(days=1)
        if cur.weekday() < 5:
            n -= 1
    return cur

def next_coupon_dates(maturity, freq_months, start, end):
    """
    從到期日往回每 freq_months 個月推一次，
    回傳所有落在 [start, end] 的配息日。
    """
    if maturity is None or not freq_months:
        return []
    out = []
    d = maturity
    # 先往回推到 end 之前（避免長天期債跑很久，直接用月差估算）
    months_back = ((maturity.year - end.year) * 12 + (maturity.month - end.month))
    k = max(0, months_back // freq_months - 1)
    d = maturity - relativedelta(months=freq_months * k)
    while d >= start:
        if d <= end:
            out.append(d)
        d -= relativedelta(months=freq_months)
    return sorted(out)

def pi_tag(b):
    """
    申購資格：
      🔒專投  = 出現在「_專投」sheet，或備註含「專業投資人」
      一般    = 出現在「非專投」sheet
      💎高資產 = 出現在「高資產」sheet（額外附註）
    回傳如 "🔒專投" / "一般" / "🔒專投💎高資產" / "未標示"
    """
    sheets = b.get("sheets") or set()
    remark = str(b.get("remark") or "")
    is_pi = any(("專投" in n and "非專投" not in n) for n in sheets) or ("專業投資人" in remark)
    is_np = any("非專投" in n for n in sheets)
    is_ha = any("高資產" in n for n in sheets)
    if is_pi:
        tag = "🔒專投"
    elif is_np:
        tag = "一般"
    else:
        tag = "未標示"
    if is_ha:
        tag += "💎高資產"
    return tag

# ---------- 讀 Excel ----------
def read_bonds(path):
    """把所有 sheet 讀成 list[dict]，並用 ISIN 去重"""
    wb = load_workbook(path, read_only=True, data_only=True)
    bonds = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        # 找標題列（含 "ISIN"）
        hdr_idx = next((i for i, r in enumerate(rows)
                        if r and any(isinstance(c, str) and "ISIN" in c for c in r)), None)
        if hdr_idx is None:
            continue
        hdr = [str(c).replace("\n", "").replace(" ", "") if c else "" for c in rows[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if "ISIN" in h: col.setdefault("isin", i)
            elif h.startswith("產品代碼"): col.setdefault("code", i)
            elif h == "債券名稱": col.setdefault("name", i)
            elif h == "幣別": col.setdefault("ccy", i)
            elif h == "債券評等": col.setdefault("sp", i)
            elif h == "債券順位": col.setdefault("seniority", i)
            elif h == "BidPrice": col.setdefault("bid", i)
            elif h == "剩餘年期": col.setdefault("years", i)
            elif h == "存續期間": col.setdefault("duration", i)
            elif h == "產品風險屬性": col.setdefault("risk", i)
            elif h == "備註": col.setdefault("remark", i)
            elif h.startswith("票面"): col.setdefault("coupon", i)
            elif h.startswith("配息"): col.setdefault("freq", i)
            elif h == "OfferPrice": col.setdefault("offer", i)
            elif h.startswith("YTM"): col.setdefault("ytm", i)
            elif h == "到期日": col.setdefault("maturity", i)
            elif h.startswith("最低申購"): col.setdefault("min_amt", i)
            elif h.startswith("本日有"): col.setdefault("avail", i)
            elif h == "交割日": col.setdefault("settle", i)
        for r in rows[hdr_idx + 2:]:          # 跳過評等副標題列
            if not r or not r[col["isin"]]:
                continue
            isin = str(r[col["isin"]]).strip()
            if not re.match(r"^[A-Z]{2}[A-Z0-9]{9}\d$", isin):
                continue
            _nm = str(r[col["name"]] if "name" in col else "").strip()
            if not _nm or _nm.startswith("#"):
                continue
            def g(k):
                return r[col[k]] if k in col and col[k] < len(r) else None
            sp_i = col.get("sp")
            ratings = ""
            if sp_i is not None:
                trio = [str(x).strip() for x in r[sp_i:sp_i+3] if x not in (None, "")]
                ratings = " / ".join(trio) if trio else ""
            b = dict(
                isin=isin, code=g("code"), name=str(g("name") or "").strip(),
                ratings=ratings, seniority=g("seniority"), bid=g("bid"),
                years=g("years"), duration=g("duration"), risk=g("risk"),
                remark=str(g("remark") or "").strip(),
                ccy=g("ccy"), coupon=g("coupon"), freq=str(g("freq") or "").strip(),
                offer=g("offer"), ytm=g("ytm"), maturity=to_date(g("maturity")),
                min_amt=g("min_amt"), avail=g("avail"), sheet=name,
                sheets=set([name]),
            )
            if isin in bonds:
                bonds[isin]["sheets"].add(name)
            else:
                bonds[isin] = b
    return list(bonds.values())

def issuer_of(name):
    """
    債券名稱 → 發行機構（歸戶用的正規化名稱）
    美林私人有限公司債3 / 美林私人公司債13 / 美林公司債6 → 美林
    高盛金融國際有限公司債25 / 高盛金融公司債39 → 高盛金融
    蘋果公司債9 → 蘋果；美國公債1 → 美國公債；西太平洋銀行債8 → 西太平洋銀行
    """
    n = re.sub(r"[\s\d０-９]+$", "", str(name)).strip()
    if n.endswith("公債"):
        return n
    if n.endswith("公司債"):
        n = n[:-3]
    elif n.endswith("債"):
        n = n[:-1]
    n = n.strip()
    # 反覆去掉法人型態字尾，讓同一集團不同發行主體歸成一家
    changed = True
    while changed:
        changed = False
        for suf in ("有限公司", "有限", "私人", "國際", "公司"):
            if n.endswith(suf) and len(n) - len(suf) >= 2:
                n = n[: -len(suf)].strip()
                changed = True
                break
    return n

def biz_days_after(d, n):
    cur = d
    while n > 0:
        cur += timedelta(days=1)
        if cur.weekday() < 5:
            n -= 1
    return cur

# ---------- 核心 ----------
def build_alerts(path, today=None, lookahead=LOOKAHEAD_DAYS):
    today = today or date.today()
    end = today + timedelta(days=lookahead)
    alerts = []
    for b in read_bonds(path):
        fm = FREQ_MONTHS.get(b["freq"])
        for cd in next_coupon_dates(b["maturity"], fm, today, end):
            last_settle = biz_days_before(cd, 1)
            lag = settle_lag(b["isin"])
            last_trade = biz_days_before(last_settle, lag)
            status = "✅ 配息前可申購" if last_trade >= today else "⛔ 本期已截止"
            alerts.append(dict(
                b, coupon_date=cd, last_settle=last_settle,
                last_trade=last_trade, lag=lag, status=status,
                days_left=(last_trade - today).days,
            ))
    alerts.sort(key=lambda a: (a["coupon_date"], -a["lag"], a["name"]))
    return alerts

def build_alert_message(path, today=None, lookahead=LOOKAHEAD_DAYS, days_ahead=3, max_lines=30, maturity_days=30):
    """
    回傳給 LINE 用的純文字訊息。
    lookahead  : 往前看幾天的配息日（預設 14）
    days_ahead : 只顯示『最晚下單日』落在今天起 N 個營業日內的（預設 3）；None = 全部顯示
    """
    today = today or date.today()
    alerts = build_alerts(path, today, lookahead)
    ok_all = [a for a in alerts if a["status"].startswith("✅")]
    gone = len(alerts) - len(ok_all)
    wd = "一二三四五六日"
    if days_ahead is None:
        ok, cutoff = ok_all, None
    else:
        cutoff = biz_days_after(today, days_ahead)
        ok = [a for a in ok_all if a["last_trade"] <= cutoff]
    scope = f"配息前申購截止日在 {days_ahead} 個營業日內（～{cutoff:%m/%d}）" if cutoff else f"未來{lookahead}天全部"
    mat_txt = format_maturities(maturing_soon(path, today, maturity_days), today, maturity_days) if maturity_days else ""
    if not ok:
        return (f"📅 {today:%m/%d}({wd[today.weekday()]}) 海外債配息雷達\n"
                f"{scope}無配息前申購截止的債券。\n"
                f"（未來{lookahead}天共 {len(alerts)} 檔配息，配息前尚可申購 {len(ok_all)} 檔，可打 /coupon all 看全部）"
                + mat_txt + DISCLAIMER)
    ok.sort(key=lambda a: (a["last_trade"], -a["lag"], a["name"]))
    lines = [f"📅 {today:%m/%d}({wd[today.weekday()]}) 海外債配息雷達",
             f"{scope}：{len(ok)} 檔",
             f"（未來{lookahead}天共 {len(alerts)} 檔配息｜配息前可申購 {len(ok_all)}｜本期已截止 {gone}）",
             "📌 配息前申購需墊付較高前手息，領回利息計入海外利息所得；配息後申購前手息較低。",
             "兩者經濟價值相當，請依客戶資金與稅務情況評估，本表僅為時點資訊。\n"]
    cur = None
    for i, a in enumerate(ok):
        if a["last_trade"] != cur:
            cur = a["last_trade"]
            tag = "⏳ 今日為配息前最後申購日" if cur == today else f"⏰ 配息前最後申購日 {cur:%m/%d}({wd[cur.weekday()]})"
            lines.append(f"── {tag} ──")
        offer = a["offer"] if a["offer"] not in (None, "", 0, "#VALUE!") else "-"
        ytm = a["ytm"] if a["ytm"] not in (None, "", 0) else "-"
        avail = "" if str(a["avail"]) == "有" else f"｜額度:{a['avail']}"
        lines.append(
            f"{a['name']} {a['ccy']} {a['coupon']}% {a['freq']}｜{pi_tag(a)}\n"
            f"  配息{a['coupon_date']:%m/%d}｜Offer {offer}｜YTM {ytm}{avail}"
        )
        if i + 1 >= max_lines and i + 1 < len(ok):
            lines.append(f"…另有 {len(ok)-i-1} 檔，見Excel")
            break
    if mat_txt:
        lines.append(mat_txt)
    lines.append(DISCLAIMER)
    return "\n".join(lines)

# ---------- 通用小工具 ----------
def first_num(v):
    """'5.32/5.33' → 5.32；'101.78' → 101.78；'-'/'#N/A'/None → None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return float(m.group()) if m else None

def find_bonds(path, keyword, max_hits=8):
    """
    找單一債券：關鍵字可含名稱片段 / ISIN / 產品代碼 / 到期年份，例如「蘋果 2043」「US037833」「蘋果9」
    回傳 list[bond]，依到期日排序
    """
    toks = [t for t in re.split(r"\s+", str(keyword).strip()) if t]
    if not toks:
        return []
    # 產品代碼縮寫：純數字（或數字片段）自動補 WMBB 前綴試比對
    # 例：26070003 → WMBB26070003；2607 → WMBB2607（片段）
    toks = ["WMBB" + t if re.fullmatch(r"\d{4,10}", t) and not re.fullmatch(r"20\d\d", t) else t for t in toks]
    years = {int(t) for t in toks if re.fullmatch(r"20\d\d", t)}
    words = [t.lower() for t in toks if not re.fullmatch(r"20\d\d", t)]
    hits = []
    for b in read_bonds(path):
        hay = f"{b['name']} {b['isin']} {b['code'] or ''}".lower().replace(" ", "")
        if all(w.replace(" ", "") in hay for w in words) and (not years or (b["maturity"] and b["maturity"].year in years)):
            hits.append(b)
    hits.sort(key=lambda b: (b["maturity"] or date.max))
    return hits[:max_hits]

# ---------- 到期日提醒 ----------
def maturing_soon(path, today=None, days=30):
    """回傳未來 days 天內到期的架上債券（依到期日排序）"""
    today = today or date.today()
    end = today + timedelta(days=days)
    out = [b for b in read_bonds(path) if b["maturity"] and today <= b["maturity"] <= end]
    out.sort(key=lambda b: (b["maturity"], b["name"]))
    return out

def format_maturities(bonds, today=None, days=30, max_lines=15):
    today = today or date.today()
    if not bonds:
        return ""
    wd = "一二三四五六日"
    lines = [f"\n💵 資金到期提醒（未來{days}天 {len(bonds)} 檔到期，本金將回流）"]
    for b in bonds[:max_lines]:
        d = b["maturity"]
        lines.append(f"▪ {d:%m/%d}({wd[d.weekday()]}) {b['name']} {b['ccy']} {b['coupon']}%｜{pi_tag(b)}")
    if len(bonds) > max_lines:
        lines.append(f"…另有 {len(bonds)-max_lines} 檔")
    lines.append("→ 建議提前聯絡持有客戶討論再投資")
    return "\n".join(lines)

# ---------- 報價檔快照與差異（信評異動 / 新上架 / 下架）----------
def bond_snapshot(path):
    """{isin: {name, ratings, ytm, offer, maturity}}"""
    snap = {}
    for b in read_bonds(path):
        snap[b["isin"]] = {
            "name": b["name"], "ratings": b.get("ratings") or "",
            "ytm": first_num(b["ytm"]), "offer": first_num(b["offer"]),
            "maturity": b["maturity"].isoformat() if b["maturity"] else None,
        }
    return snap

def diff_snapshots(old, new):
    """回傳 dict：rating_changes / new_bonds / removed_bonds"""
    old = old or {}
    rating_changes, new_bonds, removed = [], [], []
    for isin, nb in new.items():
        ob = old.get(isin)
        if ob is None:
            new_bonds.append((isin, nb["name"]))
            continue
        o_r = _norm_rating(ob.get("ratings", "")); n_r = _norm_rating(nb.get("ratings", ""))
        if o_r and n_r and o_r != n_r:
            rating_changes.append((isin, nb["name"], ob.get("ratings", ""), nb.get("ratings", "")))
    for isin, ob in old.items():
        if isin not in new:
            removed.append((isin, ob["name"]))
    return {"rating_changes": rating_changes, "new_bonds": new_bonds, "removed_bonds": removed}

def _norm_rating(r):
    """把 'N/A / Aa1 / AA+' 正規化成 'aa1/aa+'，忽略 N/A 與空白，避免假異動"""
    parts = [x.strip().lower() for x in str(r).split("/")]
    parts = [x for x in parts if x and x not in ("n/a", "na", "none", "-", "wd", "nr")]
    return "/".join(parts)

def format_snapshot_diff(diff, max_lines=15):
    lines = []
    rc, nb, rm = diff["rating_changes"], diff["new_bonds"], diff["removed_bonds"]
    if rc:
        lines.append(f"🚨 信評異動 {len(rc)} 檔（依報價檔比對）")
        for isin, name, o, n in rc[:max_lines]:
            lines.append(f"▪ {name}\n  {o} → {n}")
        if len(rc) > max_lines: lines.append(f"…另有 {len(rc)-max_lines} 檔")
    if nb:
        lines.append(f"🆕 新上架 {len(nb)} 檔")
        for isin, name in nb[:max_lines]:
            lines.append(f"▪ {name} ({isin})")
        if len(nb) > max_lines: lines.append(f"…另有 {len(nb)-max_lines} 檔")
    if rm:
        lines.append(f"📤 下架/移除 {len(rm)} 檔")
        for isin, name in rm[:max_lines]:
            lines.append(f"▪ {name}")
        if len(rm) > max_lines: lines.append(f"…另有 {len(rm)-max_lines} 檔")
    return "\n".join(lines)

# ---------- 發行機構模糊搜尋 ----------
EN_ALIAS = {
    "apple": "蘋果", "aapl": "蘋果", "microsoft": "微軟", "msft": "微軟", "amazon": "亞馬遜", "amzn": "亞馬遜",
    "google": "Alphabet", "alphabet": "Alphabet", "meta": "Meta", "nvidia": "輝達", "intel": "英特爾",
    "cisco": "思科", "oracle": "甲骨文", "ibm": "IBM", "verizon": "威瑞森", "at&t": "AT&T", "tmobile": "TMobile", "t-mobile": "TMobile",
    "merrill": "美林", "ml": "美林", "goldman": "高盛", "gs": "高盛", "morgan stanley": "摩根士丹利", "ms": "摩根士丹利",
    "jpmorgan": "摩根大通", "jpm": "摩根大通", "citi": "花旗", "citigroup": "花旗", "wells": "富國", "bofa": "美國銀行",
    "ubs": "瑞銀", "hsbc": "匯豐", "barclays": "巴克萊", "socgen": "法國興業", "societe generale": "法國興業", "bnp": "法國巴黎",
    "credit agricole": "法國農業", "deutsche": "德意志", "westpac": "西太平洋", "anz": "澳盛", "nab": "澳洲國民", "cba": "澳洲聯邦",
    "lilly": "禮來", "eli lilly": "禮來", "pfizer": "輝瑞", "abbvie": "艾伯維", "merck": "默克", "j&j": "嬌生", "johnson": "嬌生",
    "moody": "穆迪", "moodys": "穆迪", "s&p": "標普", "walmart": "沃爾瑪", "boeing": "波音", "disney": "迪士尼",
    "berkshire": "波克夏", "paypal": "Paypal", "treasury": "美國公債", "ust": "美國公債", "tsmc": "台積電",
    "vodafone": "沃達豐", "toyota": "豐田", "exxon": "埃克森", "chevron": "雪佛龍", "coca": "可口可樂", "pepsi": "百事",
    "broadcom": "博通", "qualcomm": "高通", "tesla": "特斯拉", "netflix": "Netflix", "starbucks": "星巴克", "nike": "耐吉",
}

def search_issuers(path, keyword, max_issuers=3):
    """
    在整份報價檔（不限配息中）模糊搜尋發行機構。
    比對：關鍵字（不分大小寫）包含於 發行機構名 / 債券名稱 / ISIN / 產品代碼；
    找不到時用 difflib 找最相近的機構名。
    回傳 [(issuer, [bond,...]), ...]，最多 max_issuers 家。
    """
    import difflib
    kw = str(keyword).strip().lower()
    if not kw:
        return []
    # 常見英文名 → 中文（讓 /issuer apple 也找得到）
    for en, zh in EN_ALIAS.items():
        if kw == en or (len(kw) >= 4 and (kw in en or en in kw)):
            kw = zh.lower()
            break
    groups = {}
    for b in read_bonds(path):
        iss = issuer_of(b["name"])
        b["issuer"] = iss
        groups.setdefault(iss, []).append(b)
    hits = []
    for iss, bl in groups.items():
        hay = " ".join([iss] + [b["name"] for b in bl] + [b["isin"] for b in bl] + [str(b["code"] or "") for b in bl]).lower()
        if kw in hay:
            hits.append(iss)
    if not hits:
        hits = difflib.get_close_matches(keyword, list(groups.keys()), n=max_issuers, cutoff=0.6)
    out = []
    for iss in hits[:max_issuers]:
        bl = sorted(groups[iss], key=lambda b: (b["maturity"] or date.max))
        out.append((iss, bl))
    return out

def format_issuer_bonds(issuer, bonds, intro="", max_bonds=None, today=None):
    """
    組成 LINE 文字：簡介 + 該機構架上債券摘要。
    預設全部列出（LINE 訊息由 push_long_message 自動分段）；已到期的券會被濾掉並註明檔數。
    """
    today = today or date.today()
    live = [b for b in bonds if not (b["maturity"] and b["maturity"] < today)]
    expired = len(bonds) - len(live)
    head = f"🏦 {issuer}（架上 {len(live)} 檔"
    head += f"，另 {expired} 檔已到期未列）" if expired else "）"
    lines = [head]
    if intro:
        lines.append(intro)
    lines.append("")
    show = live if max_bonds is None else live[:max_bonds]
    for b in show:
        offer = b["offer"] if b["offer"] not in (None, "", 0, "#VALUE!", "#N/A") else "-"
        ytm = b["ytm"] if b["ytm"] not in (None, "", 0, "#N/A") else "-"
        mat = f"{b['maturity']:%Y/%m/%d}" if b["maturity"] else "-"
        rt = " / ".join(x for x in str(b.get("ratings") or "").split(" / ") if x and x.upper() not in ("N/A", "NA", "NONE"))
        rating = f"｜{rt}" if rt else ""
        lines.append(f"▪ {b['name']} {b['ccy']} {b['coupon']}% {b['freq']}｜{pi_tag(b)}\n  到期{mat}｜Offer {offer}｜YTM {ytm}{rating}")
    if len(live) > len(show):
        lines.append(f"…另有 {len(live)-len(show)} 檔")
    return "\n".join(lines)

# ---------- Excel 條件表 ----------
def build_coupon_sheet(path, out_path, today=None, lookahead=LOOKAHEAD_DAYS, intro_fn=None):
    """
    把『還來得及參與』的債券輸出成單一 sheet 的 Excel 條件表（手機也看得到）。
    intro_fn(list[str]) -> dict[str,str]：可選，傳入發行機構名單，回傳簡介
    回傳 (out_path, 檔數, 發行機構數, intros_dict)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    today = today or date.today()
    alerts = [a for a in build_alerts(path, today, lookahead) if a["status"].startswith("✅")]
    alerts.sort(key=lambda a: (a["last_trade"], -a["lag"], a["name"]))
    issuers = []
    for a in alerts:
        iss = issuer_of(a["name"])
        a["issuer"] = iss
        if iss not in issuers:
            issuers.append(iss)
    intros = {}
    if intro_fn and issuers:
        try:
            intros = intro_fn(issuers) or {}
        except Exception as e:
            intros = {i: f"（簡介暫無法取得：{str(e)[:80]}）" for i in issuers}

    wb = Workbook()
    navy = PatternFill("solid", fgColor="0B2A4A")
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    bf = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D0D0D0"); bd = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws = wb.active; ws.title = "配息債條件表"
    cols = ["配息前最後申購日", "配息日", "交割", "申購資格", "發行機構", "債券名稱", "ISIN", "產品代碼", "幣別",
            "票面利率%", "配息頻率", "評等(S&P/Moody's/Fitch)", "債券順位", "Offer", "YTM/YTC",
            "到期日", "剩餘年期", "存續期間", "風險屬性", "最低申購面額", "本日額度", "備註", "發行機構簡介(AI)", "來源Sheet"]
    ws["A1"] = f"海外債配息雷達 — 配息日前可申購（{today:%Y/%m/%d} 起未來{lookahead}天，共 {len(alerts)} 檔）"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="0B2A4A")
    ws["A2"] = ("最晚交割日=配息日前1營業日；US/CA T+1、其他 T+2；營業日僅排除週六日；配息日由到期日+頻率倒推，請以實際為準。"
                "配息前申購需墊付較高前手息且利息計入海外所得，配息後申購前手息較低，兩者經濟價值相當；本表僅列時點資訊，非投資建議。"
                "發行機構簡介為 AI 產生，僅供內部參考，對客說明請以公開資訊為準。")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=4, column=j, value=c); cell.font = hf; cell.fill = navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = bd
    for i, a in enumerate(alerts, 5):
        vals = [a["last_trade"], a["coupon_date"], f"T+{a['lag']}", pi_tag(a), a["issuer"], a["name"], a["isin"], a["code"], a["ccy"],
                a["coupon"], a["freq"], a["ratings"], a["seniority"],
                a["offer"] if a["offer"] not in (None, "#VALUE!") else None,
                a["ytm"], a["maturity"], a["years"], a["duration"], a["risk"], a["min_amt"], a["avail"],
                a["remark"], intros.get(a["issuer"], ""), "、".join(sorted(a["sheets"]))]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v); cell.font = bf; cell.border = bd; cell.alignment = wrap
            if isinstance(v, date): cell.number_format = "yyyy/mm/dd"
        if a["last_trade"] == today:
            for j in range(1, len(cols) + 1):
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFF2CC")
    widths = [11, 11, 6, 12, 18, 24, 15, 15, 6, 9, 8, 20, 12, 8, 11, 11, 8, 8, 8, 12, 8, 45, 55, 28]
    for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "G5"; ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{4 + len(alerts)}"
    wb.save(out_path)
    return out_path, len(alerts), len(issuers), intros

if __name__ == "__main__":
    p = sys.argv[1]
    t = date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else date.today()
    print(build_alert_message(p, t))
